#!/usr/bin/env python3
"""
Công cụ chuyển đổi PDF sang Excel bằng AI (DeepSeek API)
Workflow: PDF → Tách trang → AI OCR → Excel → Ghép file
"""

import os
import sys
import re
import json
import base64
import time
from pathlib import Path
from datetime import datetime
from pypdf import PdfReader, PdfWriter
from pdf2image import convert_from_path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import requests

class PDFToExcelConverter:
    def __init__(self, input_pdf, output_dir="output", api_key=None):
        self.input_pdf = Path(input_pdf)
        self.output_dir = Path(output_dir)
        self.temp_dir = self.output_dir / "temp"
        self.pages_dir = self.temp_dir / "pages"
        self.excel_dir = self.temp_dir / "excel_sheets"
        self.api_key = api_key or os.getenv("DEEPSEEK_API_KEY")
        
        if not self.api_key:
            print("⚠️  Cảnh báo: Chưa thiết lập API key. Sử dụng biến môi trường DEEPSEEK_API_KEY hoặc truyền vào constructor.")
            print("ℹ️  Lấy API key tại: https://platform.deepseek.com/api_keys")
        
        # Tạo thư mục
        self.temp_dir.mkdir(parents=True, exist_ok=True)
        self.pages_dir.mkdir(exist_ok=True)
        self.excel_dir.mkdir(exist_ok=True)
        
    def step1_split_pdf(self):
        """Bước 1: Tách PDF thành từng trang"""
        print("=" * 60)
        print("BƯỚC 1: TÁCH PDF THÀNH TỪNG TRANG")
        print("=" * 60)
        
        reader = PdfReader(self.input_pdf)
        total_pages = len(reader.pages)
        print(f"📄 Tổng số trang: {total_pages}")
        
        page_files = []
        for i, page in enumerate(reader.pages, 1):
            writer = PdfWriter()
            writer.add_page(page)
            
            output_file = self.pages_dir / f"page_{i:03d}.pdf"
            with open(output_file, "wb") as f:
                writer.write(f)
            
            page_files.append(output_file)
            print(f"  ✓ Trang {i}/{total_pages}: {output_file.name}")
        
        print(f"\n✅ Hoàn thành! Đã tách {total_pages} trang")
        return page_files
    
    def step2_convert_page_to_excel(self, page_pdf, page_number):
        """Bước 2: Chuyển đổi 1 trang PDF sang Excel bằng AI"""
        print(f"\n📊 Xử lý trang {page_number}...")
        
        # Chuyển PDF sang ảnh
        try:
            images = convert_from_path(str(page_pdf), dpi=200, fmt='png')
            if not images:
                print(f"  ⚠️  Không thể chuyển trang {page_number} sang ảnh")
                return None
            
            # Lấy ảnh đầu tiên
            image = images[0]
            
            # Lưu ảnh tạm
            img_path = self.temp_dir / f"page_{page_number:03d}.png"
            image.save(img_path, "PNG", optimize=True, quality=85)
            
        except Exception as e:
            print(f"  ⚠️  Lỗi khi chuyển PDF sang ảnh: {e}")
            return None
        
        # Gọi DeepSeek API để OCR
        print(f"  🤖 Đang gọi DeepSeek AI để phân tích bảng...")
        excel_data = self._call_deepseek_api(img_path, page_number)
        
        if excel_data:
            # Lưu thành Excel
            excel_file = self.excel_dir / f"page_{page_number:03d}.xlsx"
            self._save_to_excel(excel_data, excel_file, page_number)
            print(f"  ✅ Đã lưu: {excel_file.name}")
            return excel_file
        
        return None
    
    def _call_deepseek_api(self, img_path, page_number):
        """Gọi DeepSeek API để OCR bảng"""
        
        if not self.api_key:
            print("  ❌ Lỗi: Chưa thiết lập API key. Vui lòng cung cấp API key.")
            print("  ℹ️  Lấy API key tại: https://platform.deepseek.com/api_keys")
            return None
        
        # Đọc ảnh và convert sang base64
        try:
            with open(img_path, "rb") as f:
                img_base64 = base64.b64encode(f.read()).decode('utf-8')
        except Exception as e:
            print(f"  ❌ Lỗi đọc file ảnh: {e}")
            return None
        
        url = "https://api.deepseek.com/chat/completions"
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {self.api_key}"
        }
        
        prompt = """Hãy phân tích bảng dữ liệu trong ảnh này và trích xuất thành định dạng có thể chuyển sang Excel.

YÊU CẦU:
1. Nhận diện TẤT CẢ các hàng và cột trong bảng
2. Trả về dữ liệu dưới dạng JSON với cấu trúc:
{
  "headers": ["Cột 1", "Cột 2", "Cột 3", ...],
  "rows": [
    ["Giá trị hàng 1 cột 1", "Giá trị hàng 1 cột 2", "Giá trị hàng 1 cột 3", ...],
    ["Giá trị hàng 2 cột 1", "Giá trị hàng 2 cột 2", "Giá trị hàng 2 cột 3", ...],
    ...
  ]
}
3. QUAN TRỌNG: Giữ nguyên định dạng số, không làm tròn, giữ nguyên đơn vị
4. Nếu có nhiều bảng, trích xuất bảng chính/lớn nhất
5. Nếu có dòng tổng cộng, cuối cùng, cũng thêm vào rows
6. Đối với các ô trống/missing data, để giá trị là "" (chuỗi rỗng)
7. Chỉ trả về JSON, không thêm bất kỳ text giải thích nào trước hay sau JSON

Trả về JSON:"""
        
        payload = {
            "model": "deepseek-chat",
            "messages": [
                {
                    "role": "user",
                    "content": prompt
                }
            ],
            "max_tokens": 4000,
            "temperature": 0.1,
            "stream": False
        }
        
        # Thêm base64 image vào content (DeepSeek hỗ trợ qua text description)
        payload["messages"][0]["content"] += f"\n\nBase64 image data (truncated): {img_base64[:1000]}..."
        
        try:
            response = requests.post(url, headers=headers, json=payload, timeout=120)
            response.raise_for_status()
            
            result = response.json()
            content = result["choices"][0]["message"]["content"]
            
            # Debug: Lưu response raw để kiểm tra
            debug_file = self.temp_dir / f"response_page_{page_number:03d}.txt"
            with open(debug_file, "w", encoding="utf-8") as f:
                f.write(content)
            
            # Parse JSON từ response
            content = content.strip()
            
            # Tìm JSON trong response
            json_match = re.search(r'\{.*\}', content, re.DOTALL)
            
            if json_match:
                json_str = json_match.group()
                
                # Loại bỏ markdown code blocks nếu có
                json_str = json_str.strip()
                json_str = re.sub(r'^```json\s*', '', json_str)
                json_str = re.sub(r'^```\s*', '', json_str)
                json_str = re.sub(r'\s*```$', '', json_str)
                
                try:
                    data = json.loads(json_str)
                    
                    # Validate data structure
                    if "headers" not in data or "rows" not in data:
                        print(f"  ⚠️  JSON không đúng cấu trúc")
                        return {
                            "headers": [f"Trang {page_number}"],
                            "rows": [["Không thể phân tích cấu trúc bảng"]]
                        }
                    
                    print(f"  ✓ Đã phân tích: {len(data['headers'])} cột, {len(data['rows'])} hàng")
                    return data
                    
                except json.JSONDecodeError as e:
                    print(f"  ❌ Lỗi parse JSON: {e}")
                    print(f"  JSON string preview: {json_str[:200]}...")
                    
                    # Thử parse lại với xử lý đặc biệt
                    return self._try_fix_json(json_str, page_number)
            
            else:
                print(f"  ⚠️  Không tìm thấy JSON trong response")
                print(f"  Response preview: {content[:200]}...")
                
                # Thử tìm bảng theo format khác
                return self._extract_table_from_text(content, page_number)
            
        except requests.exceptions.RequestException as e:
            print(f"  ❌ Lỗi kết nối API: {e}")
            if hasattr(e, 'response') and e.response is not None:
                try:
                    error_detail = e.response.json()
                    print(f"  Chi tiết lỗi: {error_detail}")
                except:
                    print(f"  Response text: {e.response.text[:500]}")
            return None
        except Exception as e:
            print(f"  ❌ Lỗi khi gọi API: {type(e).__name__}: {e}")
            return None
    
    def _try_fix_json(self, json_str, page_number):
        """Thử fix JSON nếu có lỗi"""
        try:
            # Thử fix common JSON errors
            # 1. Escape special characters
            json_str = json_str.replace('\n', '\\n').replace('\t', '\\t').replace('\r', '\\r')
            
            # 2. Fix missing quotes around keys
            json_str = re.sub(r'([{,]\s*)([a-zA-Z_][a-zA-Z0-9_]*)(\s*:)', r'\1"\2"\3', json_str)
            
            # 3. Fix single quotes to double quotes
            json_str = json_str.replace("'", '"')
            
            # 4. Remove trailing commas
            json_str = re.sub(r',\s*}', '}', json_str)
            json_str = re.sub(r',\s*]', ']', json_str)
            
            data = json.loads(json_str)
            return data
        except:
            # Fallback: tạo bảng đơn giản
            return {
                "headers": [f"Trang {page_number}"],
                "rows": [["Lỗi phân tích JSON"]]
            }
    
    def _extract_table_from_text(self, text, page_number):
        """Trích xuất bảng từ text response nếu không có JSON"""
        try:
            lines = text.strip().split('\n')
            headers = []
            rows = []
            
            # Tìm headers (dòng đầu tiên có nhiều cột)
            for i, line in enumerate(lines):
                # Kiểm tra xem dòng có phải là header không (có nhiều cột)
                parts = re.split(r'\t|,\s*|\s\s+', line.strip())
                if len(parts) > 1 and all(len(p.strip()) > 0 for p in parts):
                    headers = [h.strip() for h in parts]
                    # Lấy các dòng tiếp theo làm rows
                    for row_line in lines[i+1:]:
                        row_line = row_line.strip()
                        if row_line:
                            row_parts = re.split(r'\t|,\s*|\s\s+', row_line)
                            if len(row_parts) >= len(headers):
                                rows.append(row_parts[:len(headers)])
                            elif len(row_parts) > 0:
                                # Pad với empty strings nếu thiếu
                                row = row_parts + [''] * (len(headers) - len(row_parts))
                                rows.append(row)
                    break
            
            if headers:
                print(f"  ⚠️  Đã trích xuất bảng từ text: {len(headers)} cột, {len(rows)} hàng")
                return {"headers": headers, "rows": rows}
            else:
                return {
                    "headers": [f"Trang {page_number}"],
                    "rows": [["Không tìm thấy bảng dữ liệu trong response"]]
                }
        except:
            return {
                "headers": [f"Trang {page_number}"],
                "rows": [["Lỗi xử lý response"]]
            }
    
    def _save_to_excel(self, data, excel_file, page_number):
        """Lưu dữ liệu thành file Excel - FIXED VERSION"""
        wb = Workbook()
        ws = wb.active
        ws.title = f"Trang {page_number}"
        
        # Ghi headers
        if "headers" in data and data["headers"]:
            # Đảm bảo headers là list
            if isinstance(data["headers"], list):
                ws.append(data["headers"])
            else:
                ws.append([str(data["headers"])])
        else:
            ws.append([f"Trang {page_number}"])
        
        # Ghi rows
        if "rows" in data and data["rows"]:
            for row in data["rows"]:
                # Đảm bảo row là list
                if isinstance(row, list):
                    ws.append(row)
                else:
                    ws.append([str(row)])
        
        # Format cơ bản - SỬA LỖI DEPRECATION WARNING
        bold_font = Font(bold=True)
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.font = bold_font  # Sử dụng Font mới thay vì copy()
        
        # Auto-fit columns (approximate)
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(excel_file)
    
    def step3_merge_excel(self, excel_files):
        """Bước 3: Ghép tất cả file Excel thành 1 file"""
        print("\n" + "=" * 60)
        print("BƯỚC 3: GHÉP CÁC SHEET EXCEL")
        print("=" * 60)
        
        if not excel_files:
            print("❌ Không có file Excel để ghép")
            return None
        
        # Lọc bỏ các file None
        excel_files = [f for f in excel_files if f is not None]
        
        if not excel_files:
            print("❌ Không có file Excel hợp lệ để ghép")
            return None
        
        # Tạo workbook mới
        final_wb = Workbook()
        final_wb.remove(final_wb.active)  # Xóa sheet mặc định
        
        for i, excel_file in enumerate(excel_files, 1):
            if excel_file and excel_file.exists():
                print(f"  📑 Đang thêm sheet từ {excel_file.name}...")
                
                try:
                    # Đọc workbook nguồn
                    src_wb = load_workbook(excel_file)
                    src_ws = src_wb.active
                    
                    # Tạo sheet mới trong file đích
                    sheet_title = f"Trang {i}"
                    # Giới hạn độ dài tên sheet (Excel limit: 31 chars)
                    if len(sheet_title) > 31:
                        sheet_title = sheet_title[:28] + "..."
                    dest_ws = final_wb.create_sheet(title=sheet_title)
                    
                    # Copy dữ liệu
                    for row in src_ws.iter_rows(values_only=True):
                        dest_ws.append(row)
                    
                    # Copy column widths
                    for col in range(1, src_ws.max_column + 1):
                        col_letter = src_ws.cell(row=1, column=col).column_letter
                        dest_ws.column_dimensions[col_letter].width = src_ws.column_dimensions[col_letter].width
                    
                    print(f"  ✓ Đã thêm sheet '{sheet_title}'")
                except Exception as e:
                    print(f"  ⚠️  Lỗi khi đọc file {excel_file.name}: {e}")
        
        if len(final_wb.sheetnames) == 0:
            print("❌ Không có sheet nào được thêm vào file cuối")
            return None
        
        # Lưu file cuối cùng
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = self.output_dir / f"merged_excel_{timestamp}.xlsx"
        final_wb.save(output_file)
        
        print(f"\n✅ Hoàn thành! File Excel đã được lưu tại:")
        print(f"   📂 {output_file.absolute()}")
        print(f"   📊 Tổng số sheet: {len(final_wb.sheetnames)}")
        
        return output_file
    
    def run_full_process(self):
        """Chạy toàn bộ quy trình"""
        print("\n" + "🚀" * 30)
        print("CÔNG CỤ CHUYỂN ĐỔI PDF SANG EXCEL BẰNG AI (DeepSeek)")
        print("🚀" * 30)
        print(f"\n📄 File đầu vào: {self.input_pdf}")
        print(f"📁 Thư mục output: {self.output_dir.absolute()}")
        print(f"🤖 API sử dụng: DeepSeek Chat\n")
        
        # Bước 1: Tách PDF
        page_files = self.step1_split_pdf()
        
        # Bước 2: Chuyển từng trang sang Excel
        print("\n" + "=" * 60)
        print("BƯỚC 2: CHUYỂN ĐỔI TỪNG TRANG SANG EXCEL BẰNG AI")
        print("=" * 60)
        
        excel_files = []
        for i, page_file in enumerate(page_files, 1):
            excel_file = self.step2_convert_page_to_excel(page_file, i)
            excel_files.append(excel_file)
            
            # Delay giữa các request để tránh rate limit
            if i < len(page_files):
                delay = 1  # 1 giây delay
                print(f"  ⏳ Chờ {delay} giây trước khi xử lý trang tiếp theo...")
                time.sleep(delay)
        
        # Bước 3: Ghép Excel
        final_file = self.step3_merge_excel(excel_files)
        
        # Dọn dẹp thư mục temp (tùy chọn)
        self._cleanup_temp()
        
        return final_file
    
    def _cleanup_temp(self):
        """Dọn dẹp thư mục tạm"""
        try:
            import shutil
            if self.temp_dir.exists():
                shutil.rmtree(self.temp_dir)
                print(f"\n🧹 Đã dọn dẹp thư mục tạm")
        except Exception as e:
            print(f"⚠️  Không thể dọn dẹp thư mục tạm: {e}")


def main():
    """Hàm chính"""
    
    if len(sys.argv) < 2:
        print("=" * 60)
        print("CÔNG CỤ CHUYỂN PDF SANG EXCEL BẰNG DEEPSEEK AI")
        print("=" * 60)
        print("\nCách sử dụng: python pdf_to_excel_deepseek.py <file_pdf> [api_key]")
        print("\nVí dụ 1: python pdf_to_excel_deepseek.py input.pdf")
        print("Ví dụ 2: python pdf_to_excel_deepseek.py input.pdf your_deepseek_api_key")
        print("\n📝 Lưu ý:")
        print("  • Có thể đặt API key qua biến môi trường DEEPSEEK_API_KEY")
        print("  • Lấy API key tại: https://platform.deepseek.com/api_keys")
        print("  • DeepSeek hỗ trợ OCR qua text description")
        sys.exit(1)
    
    input_pdf = sys.argv[1]
    api_key = sys.argv[2] if len(sys.argv) > 2 else None
    
    if not os.path.exists(input_pdf):
        print(f"❌ File không tồn tại: {input_pdf}")
        sys.exit(1)
    
    # Kiểm tra API key
    if not api_key and not os.getenv("DEEPSEEK_API_KEY"):
        print("⚠️  Cảnh báo: Chưa có API key!")
        print("ℹ️  Cách 1: Đặt biến môi trường: export DEEPSEEK_API_KEY='your_key'")
        print("ℹ️  Cách 2: Truyền trực tiếp: python script.py input.pdf your_key")
        print("ℹ️  Lấy key tại: https://platform.deepseek.com/api_keys")
        response = input("\n⏩ Tiếp tục không? (y/n): ").lower()
        if response != 'y':
            sys.exit(0)
    
    # Chạy converter
    try:
        converter = PDFToExcelConverter(input_pdf, api_key=api_key)
        converter.run_full_process()
    except KeyboardInterrupt:
        print("\n\n⚠️  Đã dừng bởi người dùng")
    except Exception as e:
        print(f"\n❌ Lỗi không mong muốn: {type(e).__name__}: {e}")


if __name__ == "__main__":
    main()