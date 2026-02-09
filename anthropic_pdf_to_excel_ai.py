#!/usr/bin/env python3
"""
Công cụ chuyển đổi PDF sang Excel bằng AI
Workflow: PDF → Tách trang → AI OCR → Excel → Ghép file
"""

import os
import sys
from pathlib import Path
from pypdf import PdfReader, PdfWriter
from pdf2image import convert_from_path
import base64
import json
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from datetime import datetime

class PDFToExcelConverter:
    def __init__(self, input_pdf, output_dir="output"):
        self.input_pdf = Path(input_pdf)
        self.output_dir = Path(output_dir)
        self.temp_dir = self.output_dir / "temp"
        self.pages_dir = self.temp_dir / "pages"
        self.excel_dir = self.temp_dir / "excel_sheets"
        
        # Tạo thư mục
        self.temp_dir.mkdir(parents=True, exist_ok=True)
        self.pages_dir.mkdir(exist_ok=True)
        self.excel_dir.mkdir(exist_ok=True)
        
    def step1_split_pdf(self):
        """Bước 1: Tách PDF thành từng trang"""
        print("=" * 60)
        print("BƯỚC 1: TÁCH PDF THÀNH TỪNG TRANG")
        print("=" * 60)
        
        reader = PdfReader(self.input_pdf)
        total_pages = len(reader.pages)
        print(f"📄 Tổng số trang: {total_pages}")
        
        page_files = []
        for i, page in enumerate(reader.pages, 1):
            writer = PdfWriter()
            writer.add_page(page)
            
            output_file = self.pages_dir / f"page_{i:03d}.pdf"
            with open(output_file, "wb") as f:
                writer.write(f)
            
            page_files.append(output_file)
            print(f"  ✓ Trang {i}/{total_pages}: {output_file.name}")
        
        print(f"\n✅ Hoàn thành! Đã tách {total_pages} trang")
        return page_files
    
    def step2_convert_page_to_excel(self, page_pdf, page_number):
        """Bước 2: Chuyển đổi 1 trang PDF sang Excel bằng AI"""
        print(f"\n📊 Xử lý trang {page_number}...")
        
        # Chuyển PDF sang ảnh
        images = convert_from_path(page_pdf, dpi=300)
        if not images:
            print(f"  ⚠️  Không thể chuyển trang {page_number} sang ảnh")
            return None
        
        # Lấy ảnh đầu tiên
        image = images[0]
        
        # Lưu ảnh tạm
        img_path = self.temp_dir / f"page_{page_number:03d}.png"
        image.save(img_path, "PNG")
        
        # Chuyển sang base64
        with open(img_path, "rb") as f:
            img_base64 = base64.b64encode(f.read()).decode()
        
        # Gọi Claude API để OCR
        print(f"  🤖 Đang gọi AI để phân tích bảng...")
        excel_data = self._call_claude_api(img_base64, page_number)
        
        if excel_data:
            # Lưu thành Excel
            excel_file = self.excel_dir / f"page_{page_number:03d}.xlsx"
            self._save_to_excel(excel_data, excel_file, page_number)
            print(f"  ✅ Đã lưu: {excel_file.name}")
            return excel_file
        
        return None
    
    def _call_claude_api(self, img_base64, page_number):
        """Gọi Claude API để OCR bảng"""
        
        api_key = os.getenv("CLAUDE_API_KEY")
        
        url = "https://api.anthropic.com/v1/messages"
        headers = {
            "Content-Type": "application/json",
            "anthropic-version": "2023-06-01"
            "x-api-key": {api_key}
        }
        
        payload = {
            "model": "claude-sonnet-4-20250514",
            "max_tokens": 4096,
            "messages": [{
                "role": "user",
                "content": [
                    {
                        "type": "image",
                        "source": {
                            "type": "base64",
                            "media_type": "image/png",
                            "data": img_base64
                        }
                    },
                    {
                        "type": "text",
                        "text": """Hãy phân tích bảng dữ liệu trong ảnh này và trích xuất thành định dạng có thể chuyển sang Excel.

Yêu cầu:
1. Nhận diện tất cả các hàng và cột trong bảng
2. Trả về dữ liệu dưới dạng JSON với cấu trúc:
   {
     "headers": ["Cột 1", "Cột 2", ...],
     "rows": [
       ["Giá trị 1.1", "Giá trị 1.2", ...],
       ["Giá trị 2.1", "Giá trị 2.2", ...],
       ...
     ]
   }
3. Giữ nguyên định dạng số, không làm tròn
4. Nếu có nhiều bảng, trích xuất bảng chính/lớn nhất
5. Chỉ trả về JSON, không thêm text giải thích

Trả về JSON:"""
                    }
                ]
            }]
        }
        
        try:
            response = requests.post(url, headers=headers, json=payload)
            response.raise_for_status()
            
            result = response.json()
            content = result["content"][0]["text"]
            
            # Parse JSON từ response
            # Loại bỏ markdown code blocks nếu có
            content = content.strip()
            if content.startswith("```json"):
                content = content[7:]
            if content.startswith("```"):
                content = content[3:]
            if content.endswith("```"):
                content = content[:-3]
            
            data = json.loads(content.strip())
            return data
            
        except Exception as e:
            print(f"  ❌ Lỗi khi gọi API: {e}")
            return None
    
    def _save_to_excel(self, data, excel_file, page_number):
        """Lưu dữ liệu thành file Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = f"Trang {page_number}"
        
        # Ghi headers
        if "headers" in data:
            ws.append(data["headers"])
        
        # Ghi rows
        if "rows" in data:
            for row in data["rows"]:
                ws.append(row)
        
        # Format cơ bản
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.font = cell.font.copy(bold=True)
        
        wb.save(excel_file)
    
    def step3_merge_excel(self, excel_files):
        """Bước 3: Ghép tất cả file Excel thành 1 file"""
        print("\n" + "=" * 60)
        print("BƯỚC 3: GHÉP CÁC SHEET EXCEL")
        print("=" * 60)
        
        if not excel_files:
            print("❌ Không có file Excel để ghép")
            return None
        
        # Tạo workbook mới
        final_wb = Workbook()
        final_wb.remove(final_wb.active)  # Xóa sheet mặc định
        
        for i, excel_file in enumerate(excel_files, 1):
            if excel_file and excel_file.exists():
                print(f"  📑 Đang thêm sheet từ {excel_file.name}...")
                
                # Đọc workbook nguồn
                src_wb = load_workbook(excel_file)
                src_ws = src_wb.active
                
                # Tạo sheet mới trong file đích
                dest_ws = final_wb.create_sheet(title=f"Trang {i}")
                
                # Copy dữ liệu
                for row in src_ws.iter_rows():
                    dest_ws.append([cell.value for cell in row])
                
                print(f"  ✓ Đã thêm sheet 'Trang {i}'")
        
        # Lưu file cuối cùng
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = self.output_dir / f"merged_excel_{timestamp}.xlsx"
        final_wb.save(output_file)
        
        print(f"\n✅ Hoàn thành! File Excel đã được lưu tại:")
        print(f"   {output_file.absolute()}")
        
        return output_file
    
    def run_full_process(self):
        """Chạy toàn bộ quy trình"""
        print("\n" + "🚀" * 30)
        print("CÔNG CỤ CHUYỂN ĐỔI PDF SANG EXCEL BẰNG AI")
        print("🚀" * 30)
        print(f"\nFile đầu vào: {self.input_pdf}")
        print(f"Thư mục output: {self.output_dir.absolute()}\n")
        
        # Bước 1: Tách PDF
        page_files = self.step1_split_pdf()
        
        # Bước 2: Chuyển từng trang sang Excel
        print("\n" + "=" * 60)
        print("BƯỚC 2: CHUYỂN ĐỔI TỪNG TRANG SANG EXCEL BẰNG AI")
        print("=" * 60)
        
        excel_files = []
        for i, page_file in enumerate(page_files, 1):
            excel_file = self.step2_convert_page_to_excel(page_file, i)
            excel_files.append(excel_file)
            
            # Hỏi người dùng có muốn tiếp tục không
            if i < len(page_files):
                response = input(f"\n❓ Tiếp tục xử lý trang {i+1}? (y/n): ").lower()
                if response != 'y':
                    print("⏸️  Tạm dừng quá trình")
                    break
        
        # Bước 3: Ghép Excel
        final_file = self.step3_merge_excel(excel_files)
        
        return final_file


def main():
    """Hàm chính"""
    
    if len(sys.argv) < 2:
        print("Cách sử dụng: python pdf_to_excel_ai.py <file_pdf>")
        print("Ví dụ: python pdf_to_excel_ai.py input.pdf")
        sys.exit(1)
    
    input_pdf = sys.argv[1]
    
    if not os.path.exists(input_pdf):
        print(f"❌ File không tồn tại: {input_pdf}")
        sys.exit(1)
    
    # Chạy converter
    converter = PDFToExcelConverter(input_pdf)
    converter.run_full_process()


if __name__ == "__main__":
    main()
