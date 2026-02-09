#!/usr/bin/env python3
"""
Công cụ chuyển đổi PDF sang Excel bằng AI (Gemini 2.5 Flash)
Workflow: PDF -> Ảnh -> AI OCR -> Excel -> Ghép file
Cập nhật: Sử dụng SDK google-genai mới nhất
"""

import os
import sys
import re
import time
import json
from pathlib import Path
from datetime import datetime

# Thư viện xử lý PDF và Excel
from pypdf import PdfReader, PdfWriter
from pdf2image import convert_from_path
from openpyxl import Workbook, load_workbook, Font
import PIL.Image

# Thư viện Google GenAI Mới
from google import genai
from google.genai import types

class PDFToExcelConverter:
    def __init__(self, input_pdf, output_dir="output", api_key=None):
        self.input_pdf = Path(input_pdf)
        self.output_dir = Path(output_dir)
        self.temp_dir = self.output_dir / "temp"
        self.pages_dir = self.temp_dir / "pages"
        self.excel_dir = self.temp_dir / "excel_sheets"
        
        self.api_key = api_key or os.getenv("GEMINI_API_KEY")
        
        if not self.api_key:
            print("⚠️  Cảnh báo: Chưa thiết lập API key.")
            print("ℹ️  Lấy API key tại: https://aistudio.google.com/app/apikey")
            self.client = None
        else:
            # Khởi tạo Client theo SDK mới
            try:
                self.client = genai.Client(api_key=self.api_key)
            except Exception as e:
                print(f"❌ Lỗi khởi tạo Client: {e}")
                self.client = None
        
        # Tạo thư mục
        self.temp_dir.mkdir(parents=True, exist_ok=True)
        self.pages_dir.mkdir(exist_ok=True)
        self.excel_dir.mkdir(exist_ok=True)

    def step1_split_pdf(self):
        """Bước 1: Tách PDF thành từng trang"""
        print("=" * 60)
        print("BƯỚC 1: TÁCH PDF THÀNH TỪNG TRANG")
        print("=" * 60)
        
        try:
            reader = PdfReader(self.input_pdf)
            total_pages = len(reader.pages)
            print(f"📄 Tổng số trang: {total_pages}")
            
            page_files = []
            for i, page in enumerate(reader.pages, 1):
                writer = PdfWriter()
                writer.add_page(page)
                
                output_file = self.pages_dir / f"page_{i:03d}.pdf"
                with open(output_file, "wb") as f:
                    writer.write(f)
                
                page_files.append(output_file)
                print(f"  ✓ Trang {i}/{total_pages}: {output_file.name}")
            
            return page_files
        except Exception as e:
            print(f"❌ Lỗi đọc PDF: {e}")
            return []

    def step2_convert_page_to_excel(self, page_pdf, page_number):
        """Bước 2: Chuyển đổi 1 trang PDF sang Excel bằng AI"""
        print(f"\n📊 Xử lý trang {page_number}...")
        
        # Chuyển PDF sang ảnh
        try:
            images = convert_from_path(page_pdf, dpi=200, fmt='png')
            if not images:
                print(f"  ⚠️  Không thể chuyển trang {page_number} sang ảnh")
                return None
            
            # Lấy ảnh đầu tiên
            image = images[0]
            
            # Lưu ảnh tạm (để debug nếu cần)
            img_path = self.temp_dir / f"page_{page_number:03d}.png"
            image.save(img_path, "PNG")
            
        except Exception as e:
            print(f"  ⚠️  Lỗi khi chuyển PDF sang ảnh: {e}")
            print("      (Hãy chắc chắn đã cài poppler-utils)")
            return None
        
        # Gọi Gemini API
        print(f"  🤖 Đang gọi Gemini 2.5 Flash...")
        excel_data = self._call_gemini_api(image, page_number)
        
        if excel_data:
            excel_file = self.excel_dir / f"page_{page_number:03d}.xlsx"
            self._save_to_excel(excel_data, excel_file, page_number)
            print(f"  ✅ Đã lưu Excel: {excel_file.name}")
            return excel_file
        
        return None

    def _call_gemini_api(self, image_obj, page_number):
        """Gọi Gemini API bằng SDK google-genai mới"""
        if not self.client:
            return None
        
        # Sử dụng model có trong danh sách của bạn
        model_id = "gemini-2.5-flash" 

        prompt = """Trích xuất dữ liệu bảng từ hình ảnh này thành định dạng JSON.
        
        Yêu cầu bắt buộc:
        1. JSON phải có đúng cấu trúc: {"headers": ["Cột A", "Cột B"], "rows": [["Dòng 1A", "Dòng 1B"], ["Dòng 2A", "Dòng 2B"]]}
        2. Nếu có ô gộp (merged cells), hãy lặp lại giá trị hoặc xử lý sao cho hợp lý thành dạng bảng phẳng.
        3. Giữ nguyên định dạng số (ví dụ: 10,000,000) và đơn vị tiền tệ.
        4. KHÔNG thêm bất kỳ markdown (```json) nào, chỉ trả về chuỗi JSON thuần.
        """

        try:
            # Gọi API theo cú pháp mới
            response = self.client.models.generate_content(
                model=model_id,
                contents=[prompt, image_obj],
                config=types.GenerateContentConfig(
                    temperature=0.1,
                    # Hướng dẫn model trả về JSON (tính năng mới của Gemini 2.5)
                    response_mime_type="application/json" 
                )
            )
            
            if not response.text:
                print("  ⚠️ API trả về rỗng")
                return None

            json_str = response.text.strip()
            
            # Làm sạch chuỗi nếu model vẫn chèn markdown
            json_str = json_str.replace("```json", "").replace("```", "").strip()

            data = json.loads(json_str)
            
            # Validate cấu trúc
            if "headers" in data and "rows" in data:
                print(f"  ✓ Đã nhận diện: {len(data['headers'])} cột, {len(data['rows'])} dòng")
                return data
            else:
                print(f"  ⚠️ JSON thiếu trường headers hoặc rows")
                return self._fallback_data(json_str)

        except Exception as e:
            print(f"  ❌ Lỗi gọi API ({model_id}): {e}")
            return None

    def _fallback_data(self, text):
        return {
            "headers": ["Dữ liệu thô"],
            "rows": [[text[:5000]]]
        }

    def _save_to_excel(self, data, excel_file, page_number):
        wb = Workbook()
        ws = wb.active
        ws.title = f"Page {page_number}"
        
        if data.get("headers"):
            ws.append(data["headers"])
        
        if data.get("rows"):
            for row in data["rows"]:
                ws.append(row)
        
        # Format header đậm (Cách viết mới nhất)
        bold_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = bold_font
            
        wb.save(excel_file)

    def step3_merge_excel(self, excel_files):
        """Bước 3: Ghép file"""
        print("\n" + "=" * 60)
        print("BƯỚC 3: GHÉP FILE EXCEL")
        print("=" * 60)
        
        valid_files = [f for f in excel_files if f and f.exists()]
        if not valid_files:
            print("❌ Không có file để ghép")
            return None
            
        final_wb = Workbook()
        final_wb.remove(final_wb.active)
        
        for i, f in enumerate(valid_files, 1):
            try:
                src_wb = load_workbook(f)
                src_ws = src_wb.active
                
                dest_ws = final_wb.create_sheet(title=f"Trang {i}")
                for row in src_ws.iter_rows(values_only=True):
                    dest_ws.append(row)
                print(f"  ✓ Đã ghép trang {i}")
            except Exception as e:
                print(f"  ⚠️ Lỗi đọc file {f.name}: {e}")
                
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = self.output_dir / f"ket_qua_{timestamp}.xlsx"
        final_wb.save(output_file)
        
        print(f"\n✅ XONG! File lưu tại: {output_file.absolute()}")
        return output_file

    def _cleanup(self):
        import shutil
        if self.temp_dir.exists():
            shutil.rmtree(self.temp_dir)

    def run(self):
        print(f"🚀 Bắt đầu chuyển đổi: {self.input_pdf.name}")
        
        # 1. Tách trang
        pages = self.step1_split_pdf()
        
        # 2. Convert từng trang
        excel_files = []
        for i, page in enumerate(pages, 1):
            excel_files.append(self.step2_convert_page_to_excel(page, i))
            # Gemini 2.5 Flash rất nhanh và rate limit cao, 
            # nhưng ta vẫn sleep 2s để an toàn
            if i < len(pages):
                time.sleep(2)
        
        # 3. Ghép
        self.step3_merge_excel(excel_files)
        self._cleanup()

def main():
    if len(sys.argv) < 2:
        print("Sử dụng: python pdf_to_excel.py <file_pdf> [api_key]")
        sys.exit(1)
        
    pdf_path = sys.argv[1]
    key = sys.argv[2] if len(sys.argv) > 2 else None
    
    converter = PDFToExcelConverter(pdf_path, api_key=key)
    converter.run()

if __name__ == "__main__":
    main()